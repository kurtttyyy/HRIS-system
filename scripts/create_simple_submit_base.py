from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from generate_employee_dashboard_excel import employees, headers


def border(color="CBD5E1"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


wb = Workbook()
ws = wb.active
ws.title = "Employee Data Sheet"
emp = wb.create_sheet("Employee")

blue = "075985"
green = "DCFCE7"
yellow = "FEF3C7"
rose = "FFE4E6"
light_blue = "E0F2FE"
white = "FFFFFF"

for sheet in [ws, emp]:
    sheet.sheet_view.showGridLines = False
    for col in range(1, 31):
        sheet.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = 16

ws.merge_cells("B2:P3")
ws["B2"] = "Employee Data Sheet"
ws["B2"].font = Font(bold=True, size=22)
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

groups = [
    ("B5:E17", "Personal Information", rose, [("Employee ID", "D6"), ("Name", "D7"), ("Age", "D8"), ("Gender", "D9"), ("Date of Birth", "D10"), ("Marital Status", "D11"), ("Nationality", "D12"), ("NID/Passport", "D13"), ("Living Area", "D14"), ("Email", "D15")]),
    ("F5:I17", "Job Information", yellow, [("Employment Type", "H6"), ("Designation", "H7"), ("Department", "H8"), ("Section", "H9"), ("Joining Date", "H10"), ("Reporting Boss", "H11"), ("Status", "H12")]),
    ("J5:M17", "Bank & Education", light_blue, [("Bank Name", "L6"), ("Account No", "L7"), ("Payment Method", "L8"), ("Highest Education", "L11"), ("Top Training 1", "L12"), ("Top Training 2", "L13")]),
    ("N5:P17", "Salary Details", green, [("Joining Salary", "P6"), ("Current Salary", "P7"), ("Last Increment", "P8"), ("Performance Score", "P9"), ("Best Achievement", "P11")]),
]

for ref, title, fill, fields in groups:
    for row in ws[ref]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = border("E2E8F0")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    start, end = ref.split(":")
    ws[start] = title
    ws[start].font = Font(bold=True, color=blue, size=13)
    ws[start].alignment = Alignment(horizontal="center")
    for label, cell_ref in fields:
        row = ws[cell_ref].row
        col = ws[cell_ref].column
        ws.cell(row, col - 1, f"{label}:").font = Font(bold=True)
        ws.cell(row, col - 1).alignment = Alignment(horizontal="right")
        ws[cell_ref].fill = PatternFill("solid", fgColor=white)
        ws[cell_ref].border = border("94A3B8")

ws["D6"] = '=CONCATENATE("A",TEXT(COUNTA(Employee!$A$6:$A$1000)+101,"000"))'
ws["H12"] = "Active"
ws["L8"] = "Bank Transfer"

input_cells = {
    "Employee ID": "D6", "Name": "D7", "Age": "D8", "Gender": "D9", "Date of Birth": "D10",
    "Marital Status": "D11", "Nationality": "D12", "NID/Passport": "D13", "Living Area": "D14",
    "Email": "D15", "Employment Type": "H6", "Designation": "H7", "Department": "H8", "Section": "H9",
    "Joining Date": "H10", "Reporting Boss": "H11", "Status": "H12", "Bank Name": "L6",
    "Account No": "L7", "Payment Method": "L8", "Highest Education": "L11", "Top Training 1": "L12",
    "Top Training 2": "L13", "Joining Salary": "P6", "Current Salary": "P7", "Last Increment": "P8",
    "Performance Score": "P9", "Best Achievement": "P11",
}

ws.merge_cells("B20:P20")
ws["B20"] = "Generated row for Employee page"
ws["B20"].font = Font(bold=True, color=white, size=14)
ws["B20"].fill = PatternFill("solid", fgColor=blue)
ws["B20"].alignment = Alignment(horizontal="center")

for col, header in enumerate(headers, 1):
    h = ws.cell(22, col, header)
    h.fill = PatternFill("solid", fgColor=blue)
    h.font = Font(bold=True, color=white, size=9)
    h.border = border("0E7490")
    h.alignment = Alignment(horizontal="center", wrap_text=True)
    c = ws.cell(23, col)
    if header == "Initials":
        c.value = '=IF(D7="","",UPPER(LEFT(D7,1)&IFERROR(MID(D7,FIND(" ",D7&" ")+1,1),"")))'
    elif header == "Service Age":
        c.value = '=IF(H10="","",ROUND((TODAY()-H10)/365.25,1))'
    elif header in input_cells:
        c.value = f"={input_cells[header]}"
    c.border = border()

ws.merge_cells("N18:P19")
ws["N18"] = "Submit button will be added in macro-enabled version"
ws["N18"].fill = PatternFill("solid", fgColor="0F766E")
ws["N18"].font = Font(bold=True, color=white)
ws["N18"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col, header in enumerate(headers, 1):
    c = emp.cell(5, col, header)
    c.fill = PatternFill("solid", fgColor=blue)
    c.font = Font(bold=True, color=white, size=9)
    c.border = border("0E7490")
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for row_index, employee in enumerate(employees, 6):
    for col, header in enumerate(headers, 1):
        c = emp.cell(row_index, col, employee[header])
        c.border = border()
        c.alignment = Alignment(vertical="center", wrap_text=True)

wb.active = 0
wb.save("employee_submit_simple_base.xlsx")
print("employee_submit_simple_base.xlsx")
