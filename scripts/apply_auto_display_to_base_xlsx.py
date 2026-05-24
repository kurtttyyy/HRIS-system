from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border


path = Path("employee_submit_enabled_base.xlsx")
backup = Path("employee_submit_enabled_base.backup.xlsx")

if not backup.exists():
    copy2(path, backup)

wb = load_workbook(path)
ws = wb["Employee"]

side = Side(style="thin", color="CBD5E1")
border = Border(left=side, right=side, top=side, bottom=side)

target_row = 12

for col in range(1, 31):
    cell = ws.cell(target_row, col)
    cell.value = f"='Employee Data Sheet'!{ws.cell(23, col).coordinate}"
    cell.border = border
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor="ECFDF5")

note_row = target_row + 2
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=10)
note = ws.cell(note_row, 1)
note.value = "Auto display row: this row shows the current data entered on Employee Data Sheet. To permanently keep many employees in .xlsx, copy this row and paste values into the next blank row."
note.font = Font(italic=True, color="334155")
note.fill = PatternFill("solid", fgColor="E0F2FE")
note.border = border
note.alignment = Alignment(vertical="center", wrap_text=True)

wb.save(path)
print(path.resolve())
print(backup.resolve())
