from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


INPUT = Path("employee_dashboard_cover_beautiful_employee_headers_20260513_165320.xlsx")
OUTPUT = Path(f"employee_dashboard_cover_beautiful_final_data_entry_{datetime.now():%Y%m%d_%H%M%S}.xlsx")

GREEN = "66F2C6"
BLUE = "00AEEF"
YELLOW = "FFF200"
ORANGE = "F79646"
PALE_GREEN = "D9FBEA"
PALE_BLUE = "DDF3FF"
PALE_YELLOW = "FFF9C4"
PALE_ORANGE = "FCE4D6"
WHITE = "FFFFFF"
BLACK = "000000"
DARK = "0F172A"
NAV_BLUE = "075985"
NAV_TEAL = "0F766E"

HIDDEN_START = 80
EMPLOYEE_ROW_START = 6
ENTRY_ROWS = 100


FIELDS = [
    ("NAME", GREEN, "Personal Information"),
    ("ID Number", GREEN, "Personal Information"),
    ("Account #", GREEN, "Personal Information"),
    ("SEX", GREEN, "Personal Information"),
    ("CIVIL STATUS", GREEN, "Personal Information"),
    ("ADDRESS", GREEN, "Personal Information"),
    ("CONTACT NO.", GREEN, "Personal Information"),
    ("DATE OF BIRTH", GREEN, "Personal Information"),
    ("AGE TO DATE", GREEN, "Personal Information"),
    ("EMPLOYMENT DATE", GREEN, "Employment Information"),
    ("LENGTH OF SERVICE", GREEN, "Employment Information"),
    ("POSITION", GREEN, "Employment Information"),
    ("DEPARTMENT", GREEN, "Employment Information"),
    ("CLASS", GREEN, "Employment Information"),
    ("RANK", GREEN, "Employment Information"),
    ("GRADE", GREEN, "Employment Information"),
    ("SSS", BLUE, "Government Numbers"),
    ("TIN", BLUE, "Government Numbers"),
    ("PHILHEALTH", BLUE, "Government Numbers"),
    ("PAG-IBIG MID", BLUE, "Government Numbers"),
    ("PAG-IBIG RTN", BLUE, "Government Numbers"),
    ("SCHOOL NAME", YELLOW, "Education"),
    ("BACHELOR", YELLOW, "Education"),
    ("MASTER'S DEGREE", YELLOW, "Education"),
    ("DOCTORATE DEGREE", YELLOW, "Education"),
    ("With/Without License", YELLOW, "License / Eligibility"),
    ("ELIGIBILITY", YELLOW, "License / Eligibility"),
    ("Registration No.", YELLOW, "License / Eligibility"),
    ("Registration Date", YELLOW, "License / Eligibility"),
    ("Valid Until", YELLOW, "License / Eligibility"),
    ("Rate per Hour", GREEN, "Salary Details"),
    ("Basic Salary", GREEN, "Salary Details"),
    ("Allowance", GREEN, "Salary Details"),
    ("Date Resigned", ORANGE, "Exit / History"),
    ("Employment History", ORANGE, "Exit / History"),
]

SECTION_FILLS = {
    "Personal Information": PALE_GREEN,
    "Employment Information": PALE_GREEN,
    "Government Numbers": PALE_BLUE,
    "Education": PALE_YELLOW,
    "License / Eligibility": PALE_YELLOW,
    "Salary Details": PALE_GREEN,
    "Exit / History": PALE_ORANGE,
}


def thin(color="BFBFBF"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def add_nav(ws, active):
    items = [
        ("Cover", "Cover", "A3"),
        ("Dashboard", "Dashboard", "A5"),
        ("Summary", "Summary", "A7"),
        ("Data Entry", "Data Entry", "A9"),
        ("Employee", "Employee", "A11"),
    ]
    ws.column_dimensions["A"].width = 14
    for label, target, cell_ref in items:
        cell = ws[cell_ref]
        cell.value = label
        sheet_ref = f"'{target}'" if " " in target else target
        cell.hyperlink = f"#{sheet_ref}!A1"
        cell.fill = PatternFill("solid", fgColor=NAV_TEAL if target == active else NAV_BLUE)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin("0E7490")


def clear_sheet(ws):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for row in range(1, max(ws.max_row, 180) + 1):
        for col in range(1, max(ws.max_column, 24) + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.hyperlink = None
            cell.fill = PatternFill("solid", fgColor=WHITE)
            cell.border = Border()
            cell.font = Font(color=BLACK, size=10)
            cell.alignment = Alignment(vertical="center")


def draw_section(ws, title, start_row, start_col, fields):
    fill = SECTION_FILLS[title]
    panel_width = 7
    end_col = start_col + panel_width - 1
    end_row = start_row + len(fields) + 1

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = thin("D9E2F3")

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(start_row, start_col)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=11, color=NAV_BLUE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    input_cells = {}
    for offset, (field, color, _) in enumerate(fields, start=start_row + 1):
        ws.merge_cells(start_row=offset, start_column=start_col, end_row=offset, end_column=start_col + 2)
        label = ws.cell(offset, start_col)
        label.value = f"{field}:"
        label.font = Font(bold=True, size=9, color=BLACK)
        label.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)

        ws.merge_cells(start_row=offset, start_column=start_col + 3, end_row=offset, end_column=end_col)
        value = ws.cell(offset, start_col + 3)
        value.fill = PatternFill("solid", fgColor=WHITE)
        value.border = thin("8EAADB")
        value.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        input_cells[field] = value.coordinate

    return input_cells


def add_validation(ws, cell_ref, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws[cell_ref])


def rebuild_data_entry(wb):
    if "Data Entry" in wb.sheetnames:
        ws = wb["Data Entry"]
    else:
        ws = wb.create_sheet("Data Entry", 3)

    clear_sheet(ws)
    add_nav(ws, "Data Entry")
    ws.sheet_view.showGridLines = False

    for col in range(2, 17):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for col in range(17, 25):
        ws.column_dimensions[get_column_letter(col)].width = 4
    for row in range(1, 76):
        ws.row_dimensions[row].height = 20

    ws.merge_cells("B2:P3")
    title = ws["B2"]
    title.value = "Employee Data Entry"
    title.font = Font(name="Aptos Display", bold=True, size=22, color=DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")

    sections = {
        "Personal Information": (5, 2),
        "Employment Information": (5, 10),
        "Government Numbers": (18, 2),
        "Education": (18, 10),
        "License / Eligibility": (27, 2),
        "Salary Details": (27, 10),
        "Exit / History": (38, 2),
    }

    input_cells = {}
    for section, (row, col) in sections.items():
        section_fields = [field for field in FIELDS if field[2] == section]
        input_cells.update(draw_section(ws, section, row, col, section_fields))

    add_validation(ws, input_cells["SEX"], '"Male,Female,Other"')
    add_validation(ws, input_cells["CIVIL STATUS"], '"Single,Married,Widowed,Separated"')
    add_validation(ws, input_cells["With/Without License"], '"With License,Without License"')
    for field in ["DATE OF BIRTH", "EMPLOYMENT DATE", "Registration Date", "Valid Until", "Date Resigned"]:
        ws[input_cells[field]].number_format = "d-mmm-yyyy"
    for field in ["Rate per Hour", "Basic Salary", "Allowance"]:
        ws[input_cells[field]].number_format = '#,##0.00'

    ws[input_cells["AGE TO DATE"]] = f'=IF({input_cells["DATE OF BIRTH"]}="","",DATEDIF({input_cells["DATE OF BIRTH"]},TODAY(),"Y"))'
    ws[input_cells["LENGTH OF SERVICE"]] = f'=IF({input_cells["EMPLOYMENT DATE"]}="","",DATEDIF({input_cells["EMPLOYMENT DATE"]},TODAY(),"Y")&" year(s)")'

    ws.merge_cells("B47:L49")
    note = ws["B47"]
    note.value = "Fill the fields above, then save the workbook. This entry appears on the Employee sheet in the first row."
    note.fill = PatternFill("solid", fgColor="E0F2FE")
    note.font = Font(italic=True, color="334155", size=10)
    note.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    note.border = thin("7DD3FC")

    ws.merge_cells("N47:P49")
    submit = ws["N47"]
    submit.value = "Submit"
    submit.fill = PatternFill("solid", fgColor="064E3B")
    submit.font = Font(bold=True, color=WHITE, size=14)
    submit.alignment = Alignment(horizontal="center", vertical="center")
    submit.border = Border(
        left=Side(style="medium", color="B45309"),
        right=Side(style="medium", color="B45309"),
        top=Side(style="medium", color="B45309"),
        bottom=Side(style="medium", color="B45309"),
    )
    submit.hyperlink = "#Employee!B6"

    for idx, (field, color, _) in enumerate(FIELDS, start=2):
        header = ws.cell(HIDDEN_START, idx)
        header.value = field
        header.fill = PatternFill("solid", fgColor=color)
        header.font = Font(bold=True, size=10, color=BLACK)
        header.alignment = Alignment(horizontal="center", vertical="center")
        header.border = thin("000000")

        value = ws.cell(HIDDEN_START + 1, idx)
        value.value = f"={input_cells[field]}"
        value.border = thin()

    for row in range(HIDDEN_START + 2, HIDDEN_START + ENTRY_ROWS):
        for col in range(2, 2 + len(FIELDS)):
            ws.cell(row, col).value = ""
            ws.cell(row, col).border = thin()

    for row in range(HIDDEN_START, HIDDEN_START + ENTRY_ROWS):
        ws.row_dimensions[row].hidden = True

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 90
    return input_cells


def rebuild_employee_links(wb):
    ws = wb["Employee"]
    header_count = len(FIELDS)

    for row in range(6, 106):
        for col in range(2, 2 + header_count):
            target = ws.cell(row, col)
            source_col = get_column_letter(col)
            source_row = HIDDEN_START + 1 + (row - EMPLOYEE_ROW_START)
            target.value = f'=IF(\'Data Entry\'!$B{source_row}="","",\'Data Entry\'!{source_col}{source_row})'
            target.fill = PatternFill("solid", fgColor=WHITE)
            target.border = thin("BFBFBF")
            target.alignment = Alignment(vertical="center", wrap_text=True)


def main():
    wb = load_workbook(INPUT)
    rebuild_data_entry(wb)
    rebuild_employee_links(wb)
    wb.save(OUTPUT)
    print(OUTPUT.resolve())


if __name__ == "__main__":
    main()
