from copy import copy
from datetime import datetime
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


WORKBOOK = Path("employee_dashboard_cover_beautiful.xlsx")
BACKUP = Path("employee_dashboard_cover_beautiful.before-data-entry.xlsx")
FALLBACK_OUTPUT = Path("employee_dashboard_cover_beautiful_with_data_entry.xlsx")
ENTRY_SHEET = "Data Entry"
EMPLOYEE_SHEET = "Employee"
FIRST_EMPLOYEE_ROW = 6
EXISTING_EMPLOYEE_ROWS = 0
FIRST_ENTRY_ROW = 6
ENTRY_ROWS = 100


BLUE = "075985"
TEAL = "0F766E"
DARK = "0F172A"
WHITE = "FFFFFF"
GRAY = "F8FAFC"
LINE = "CBD5E1"
PALE_GREEN = "ECFDF5"
PALE_BLUE = "E0F2FE"
PALE_YELLOW = "FEF3C7"
PALE_ROSE = "FFE4E6"
PALE_SECTION_BLUE = "DFF1FB"


def thin(color=LINE):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def clone_cell_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def reset_sheet(wb, title):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title, 3)


def add_nav(ws, active):
    items = [
        ("Cover", "Cover", "A3"),
        ("Dashboard", "Dashboard", "A5"),
        ("Summary", "Summary", "A7"),
        ("Data Entry", ENTRY_SHEET, "A9"),
        ("Employee", EMPLOYEE_SHEET, "A11"),
    ]
    ws.column_dimensions["A"].width = 14
    for label, target, cell_ref in items:
        cell = ws[cell_ref]
        cell.value = label
        sheet_ref = f"'{target}'" if " " in target else target
        cell.hyperlink = f"#{sheet_ref}!A1"
        cell.fill = PatternFill("solid", fgColor=TEAL if target == active else BLUE)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin("0E7490")
        ws.row_dimensions[cell.row].height = 26


def create_entry_sheet(wb, headers, employee_ws):
    ws = reset_sheet(wb, ENTRY_SHEET)
    ws.sheet_view.showGridLines = False
    add_nav(ws, ENTRY_SHEET)

    for row in ws["B1:AD140"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=WHITE)
            cell.border = Border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells("B2:P3")
    title = ws["B2"]
    title.value = "Employee Data Entry"
    title.font = Font(name="Aptos Display", bold=True, size=24, color=DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Shape the visible input form to match the supplied reference image.
    column_widths = {
        "B": 3,
        "C": 11,
        "D": 18,
        "E": 14,
        "F": 3,
        "G": 3,
        "H": 3,
        "I": 4,
        "J": 3,
        "K": 11,
        "L": 18,
        "M": 14,
        "N": 3,
        "O": 3,
        "P": 3,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    for col in range(17, 31):
        ws.column_dimensions[get_column_letter(col)].width = 4
    for row in range(1, 46):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[38].height = 30

    input_cells = {
        "Employee ID": "D6",
        "Name": "D7",
        "Age": "D8",
        "Gender": "D9",
        "Date of Birth": "D10",
        "Marital Status": "D11",
        "Nationality": "D12",
        "NID/Passport": "D13",
        "Living Area": "D14",
        "Email": "D15",
        "Employment Type": "L6",
        "Designation": "L7",
        "Department": "L8",
        "Section": "L9",
        "Joining Date": "L10",
        "Reporting Boss": "L11",
        "Status": "L12",
        "Bank Name": "D24",
        "Account No": "D25",
        "Payment Method": "D26",
        "Highest Education": "D29",
        "Top Training 1": "D30",
        "Top Training 2": "D31",
        "Joining Salary": "L24",
        "Current Salary": "L25",
        "Last Increment": "L26",
        "Performance Score": "L27",
        "Best Achievement": "L29",
    }

    sections = [
        ("B5:H19", "Personal Information", PALE_ROSE, "B5", [
            ("Employee ID", "D6"),
            ("Name", "D7"),
            ("Age", "D8"),
            ("Gender", "D9"),
            ("Date of Birth", "D10"),
            ("Marital Status", "D11"),
            ("Nationality", "D12"),
            ("NID/Passport", "D13"),
            ("Living Area", "D14"),
            ("Email", "D15"),
        ]),
        ("J5:P19", "Job Information", PALE_YELLOW, "J5", [
            ("Employment Type", "L6"),
            ("Designation", "L7"),
            ("Department", "L8"),
            ("Section", "L9"),
            ("Joining Date", "L10"),
            ("Reporting Boss", "L11"),
            ("Status", "L12"),
        ]),
        ("B23:H35", "Bank & Education", PALE_SECTION_BLUE, "B23", [
            ("Bank Name", "D24"),
            ("Account No", "D25"),
            ("Payment Method", "D26"),
            ("Highest Education", "D29"),
            ("Top Training 1", "D30"),
            ("Top Training 2", "D31"),
        ]),
        ("J23:P35", "Salary Details", "DCFCE7", "J23", [
            ("Joining Salary", "L24"),
            ("Current Salary", "L25"),
            ("Last Increment", "L26"),
            ("Performance Score", "L27"),
            ("Best Achievement", "L29"),
        ]),
    ]

    for ref, title_text, fill, title_cell_ref, fields in sections:
        for row in ws[ref]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.border = thin("DCE6EF")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        section_title = ws[title_cell_ref]
        section_title.value = title_text
        section_title.font = Font(bold=True, size=11, color=BLUE)
        section_title.alignment = Alignment(horizontal="left", vertical="center")

        for label, cell_ref in fields:
            row_num = ws[cell_ref].row
            col_num = ws[cell_ref].column
            label_start_col = col_num - 2
            label_end_col = col_num - 1
            if label_start_col >= 1:
                ws.merge_cells(
                    start_row=row_num,
                    start_column=label_start_col,
                    end_row=row_num,
                    end_column=label_end_col,
                )
            label_cell = ws.cell(row_num, label_start_col)
            label_cell.value = f"{label}:"
            label_cell.font = Font(bold=True, size=9, color="000000")
            label_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)

            value_cell = ws[cell_ref]
            value_cell.fill = PatternFill("solid", fgColor=WHITE)
            value_cell.border = thin("8EAADB")
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["D6"] = ""
    ws["L12"] = "Active"
    ws["D26"] = "Bank Transfer"

    for cell_ref in ["D10", "L10"]:
        ws[cell_ref].number_format = "d-mmm-yyyy"
    for cell_ref in ["L24", "L25", "L26"]:
        ws[cell_ref].number_format = '#,##0'

    validations = {
        "D9": '"Male,Female,Other"',
        "D11": '"Single,Married,Widowed,Separated"',
        "L6": '"Permanent,Contract,Temporary,Intern"',
        "L8": '"Production,Quality,HR,IT,Accounts"',
        "L12": '"Active,Inactive"',
        "D26": '"Bank Transfer,Cash,Check"',
    }
    for cell_ref, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws[cell_ref])

    ws.merge_cells("N37:P39")
    submit = ws["N37"]
    submit.value = "Submit"
    submit.fill = PatternFill("solid", fgColor="064E3B")
    submit.font = Font(bold=True, size=14, color=WHITE)
    submit.alignment = Alignment(horizontal="center", vertical="center")
    submit.border = Border(
        left=Side(style="medium", color="B45309"),
        right=Side(style="medium", color="B45309"),
        top=Side(style="medium", color="B45309"),
        bottom=Side(style="medium", color="B45309"),
    )
    submit.hyperlink = f"#{EMPLOYEE_SHEET}!A12"

    ws.merge_cells("B37:M39")
    helper = ws["B37"]
    helper.value = "After filling the form, save the workbook. The entered employee appears on the Employee page in the first row."
    helper.fill = PatternFill("solid", fgColor=PALE_BLUE)
    helper.font = Font(italic=True, size=10, color="334155")
    helper.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    helper.border = thin("7DD3FC")

    # Hidden normalized table. Employee sheet links to this row block, so the visible
    # form behaves like a save/submit form without requiring VBA macros.
    hidden_start = 50
    for col, header in enumerate(headers, 1):
        header_cell = ws.cell(hidden_start, col, header)
        header_cell.fill = PatternFill("solid", fgColor=BLUE)
        header_cell.font = Font(bold=True, color=WHITE, size=9)
        header_cell.border = thin("0E7490")
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        value_cell = ws.cell(hidden_start + 1, col)
        source = input_cells.get(header)
        if header == "Initials":
            value_cell.value = '=IF(D7="","",UPPER(LEFT(D7,1)&IFERROR(MID(D7,FIND(" ",D7&" ")+1,1),"")))'
        elif header == "Service Age":
            value_cell.value = '=IF(L10="","",ROUND((TODAY()-L10)/365.25,1))'
        elif source:
            value_cell.value = f"={source}"
        else:
            value_cell.value = ""
        value_cell.border = thin("E2E8F0")

    for row in range(hidden_start + 2, hidden_start + ENTRY_ROWS):
        for col in range(1, 31):
            ws.cell(row, col).value = ""
            ws.cell(row, col).fill = PatternFill("solid", fgColor=WHITE)
            ws.cell(row, col).border = thin("E2E8F0")

    table_ref = f"A{hidden_start}:AD{hidden_start + ENTRY_ROWS - 1}"
    tab = Table(displayName="NewEmployeeEntries", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    for row_num in range(hidden_start, hidden_start + ENTRY_ROWS):
        ws.row_dimensions[row_num].hidden = True

    ws.freeze_panes = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 85
    return ws


def link_entry_rows_to_employee(employee_ws, headers):
    first_linked_employee_row = FIRST_EMPLOYEE_ROW + EXISTING_EMPLOYEE_ROWS
    last_linked_employee_row = first_linked_employee_row + ENTRY_ROWS - 1

    for row in range(first_linked_employee_row, last_linked_employee_row + 1):
        entry_row = 51 + (row - first_linked_employee_row)
        for col, header in enumerate(headers, 1):
            cell = employee_ws.cell(row, col)
            source_style = employee_ws.cell(FIRST_EMPLOYEE_ROW, col)
            clone_cell_style(source_style, cell)
            cell.fill = PatternFill("solid", fgColor=PALE_GREEN if row % 2 else WHITE)
            cell.border = thin("D1FAE5")
            letter = get_column_letter(col)
            if header == "Initials":
                cell.value = f'=IF(\'{ENTRY_SHEET}\'!$A{entry_row}="","",\'{ENTRY_SHEET}\'!C{entry_row})'
            elif header == "Service Age":
                cell.value = f'=IF(\'{ENTRY_SHEET}\'!$A{entry_row}="","",\'{ENTRY_SHEET}\'!Q{entry_row})'
            else:
                cell.value = f'=IF(\'{ENTRY_SHEET}\'!$A{entry_row}="","",\'{ENTRY_SHEET}\'!{letter}{entry_row})'

    table = employee_ws.tables.get("EmployeeTable")
    if table:
        table.ref = f"A5:AD{last_linked_employee_row}"

    employee_ws.auto_filter.ref = f"A5:AD{last_linked_employee_row}"
    employee_ws.freeze_panes = "A6"


def update_navs(wb):
    for sheet_name in ["Dashboard", "Summary"]:
        if sheet_name in wb.sheetnames:
            add_nav(wb[sheet_name], sheet_name)


def update_formulas(wb):
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = cell.value.replace("$105", "$105").replace("$111", "$105")

        for dv in ws.data_validations.dataValidation:
            if dv.formula1:
                dv.formula1 = dv.formula1.replace("$111", "$105")

        if ws["N3"].value == "A101":
            ws["N3"] = ""

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        ws["C4"] = '=COUNTIF(Employee!$E$6:$E$105,"Male")'
        ws["C5"] = '=COUNTIF(Employee!$E$6:$E$105,"Female")'
        ws["C6"] = '=COUNTIF(Employee!$E$6:$E$105,"Other")'
        ws["C8"] = '=COUNTIF(Employee!$AC$6:$AC$105,"Active")'
        ws["C9"] = '=IFERROR(AVERAGEIF(Employee!$AC$6:$AC$105,"Active",Employee!$Z$6:$Z$105),0)'
        ws["C10"] = '=IFERROR(AVERAGEIF(Employee!$A$6:$A$105,"?*",Employee!$AB$6:$AB$105),0)'
        for row in range(4, 9):
            ws.cell(row, 6).value = f'=COUNTIF(Employee!$N$6:$N$105,E{row})'
        for row in range(4, 10):
            ws.cell(row, 9).value = f'=SUMPRODUCT(--(Employee!$A$6:$A$105<>""),--(YEAR(Employee!$P$6:$P$105)=H{row}))'

        ws.conditional_formatting.add(
            "L4:L103",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color="2DD4BF"),
        )


def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)
    if not BACKUP.exists():
        copy2(WORKBOOK, BACKUP)

    wb = load_workbook(WORKBOOK)
    employee_ws = wb[EMPLOYEE_SHEET]
    headers = [employee_ws.cell(5, col).value for col in range(1, 31)]

    create_entry_sheet(wb, headers, employee_ws)
    link_entry_rows_to_employee(employee_ws, headers)
    update_navs(wb)
    update_formulas(wb)

    wb._sheets = [wb["Cover"], wb["Dashboard"], wb["Summary"], wb[ENTRY_SHEET], wb[EMPLOYEE_SHEET]]
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    try:
        wb.save(WORKBOOK)
        output = WORKBOOK
    except PermissionError:
        try:
            wb.save(FALLBACK_OUTPUT)
            output = FALLBACK_OUTPUT
        except PermissionError:
            output = Path(f"employee_dashboard_cover_beautiful_with_data_entry_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
            wb.save(output)

    print(output.resolve())
    print(BACKUP.resolve())


if __name__ == "__main__":
    main()
