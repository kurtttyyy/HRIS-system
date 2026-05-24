from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT = Path("employee_dashboard_cover_beautiful_no_people.xlsx")
FALLBACK_INPUT = Path("employee_dashboard_cover_beautiful.xlsx")
OUTPUT = Path(f"employee_dashboard_cover_beautiful_employee_headers_{datetime.now():%Y%m%d_%H%M%S}.xlsx")

GREEN = "66F2C6"
BLUE = "00AEEF"
YELLOW = "FFF200"
ORANGE = "F79646"
WHITE = "FFFFFF"
BLACK = "000000"
DARK_BLUE = "075985"
TEAL = "0F766E"


def thin(color="000000"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def add_nav(ws):
    items = [
        ("Cover", "Cover", "A3"),
        ("Dashboard", "Dashboard", "A5"),
        ("Summary", "Summary", "A7"),
        ("Data Entry", "Data Entry", "A9"),
        ("Employee", "Employee", "A11"),
    ]
    ws.column_dimensions["A"].width = 14
    for label, target, cell_ref in items:
        cell = ws[cell_ref]
        cell.value = label
        sheet_ref = f"'{target}'" if " " in target else target
        cell.hyperlink = f"#{sheet_ref}!A1"
        cell.fill = PatternFill("solid", fgColor=TEAL if target == "Employee" else DARK_BLUE)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin("0E7490")


def set_band(ws, row, start_col, end_col, text, fill):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col)
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=BLACK, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = thin()
    for col in range(start_col, end_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row, col).border = thin()


def style_data_area(ws, row_start=9, row_end=60, col_start=2, col_end=40):
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=WHITE)
            cell.border = thin("BFBFBF")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22


def main():
    source = INPUT if INPUT.exists() else FALLBACK_INPUT
    wb = load_workbook(source)
    ws = wb["Employee"]

    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]

    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    for row in range(1, ws.max_row + 1):
        for col in range(1, max(ws.max_column, 40) + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = PatternFill("solid", fgColor=WHITE)
            cell.border = Border()
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(color=BLACK, size=10)

    add_nav(ws)

    ws.merge_cells("B2:AN3")
    title = ws["B2"]
    title.value = "Employee Data Registry"
    title.font = Font(name="Aptos Display", bold=True, size=20, color=BLACK)
    title.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(2, 41):
        ws.column_dimensions[get_column_letter(col)].width = 10

    widths = {
        "B": 13, "C": 13, "D": 10, "E": 10, "F": 10, "G": 11, "H": 11,
        "I": 8, "J": 10, "K": 10, "L": 13, "M": 13, "N": 13, "O": 12,
        "P": 12, "Q": 12, "R": 12, "S": 12, "T": 12, "U": 12, "V": 12,
        "W": 12, "X": 12, "Y": 12, "Z": 12, "AA": 12, "AB": 12,
        "AC": 12, "AD": 12, "AE": 12, "AF": 12, "AG": 12, "AH": 12,
        "AI": 12, "AJ": 12, "AK": 12, "AL": 12, "AM": 12, "AN": 12,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(5, 9):
        ws.row_dimensions[row].height = 22

    one_row_headers = [
        ("NAME", GREEN),
        ("ID Number", GREEN),
        ("Account #", GREEN),
        ("SEX", GREEN),
        ("CIVIL STATUS", GREEN),
        ("ADDRESS", GREEN),
        ("CONTACT NO.", GREEN),
        ("DATE OF BIRTH", GREEN),
        ("AGE TO DATE", GREEN),
        ("EMPLOYMENT DATE", GREEN),
        ("LENGTH OF SERVICE", GREEN),
        ("POSITION", GREEN),
        ("DEPARTMENT", GREEN),
        ("CLASS", GREEN),
        ("RANK", GREEN),
        ("GRADE", GREEN),
        ("SSS", BLUE),
        ("TIN", BLUE),
        ("PHILHEALTH", BLUE),
        ("PAG-IBIG MID", BLUE),
        ("PAG-IBIG RTN", BLUE),
        ("SCHOOL NAME", YELLOW),
        ("BACHELOR", YELLOW),
        ("MASTER'S DEGREE", YELLOW),
        ("DOCTORATE DEGREE", YELLOW),
        ("With/Without License", YELLOW),
        ("ELIGIBILITY", YELLOW),
        ("Registration No.", YELLOW),
        ("Registration Date", YELLOW),
        ("Valid Until", YELLOW),
        ("Rate per Hour", GREEN),
        ("Basic Salary", GREEN),
        ("Allowance", GREEN),
        ("Date Resigned", ORANGE),
        ("Employment History", ORANGE),
    ]

    header_row = 5
    for offset, (text, fill) in enumerate(one_row_headers, start=2):
        cell = ws.cell(header_row, offset)
        cell.value = text
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color=BLACK, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin()
        ws.column_dimensions[get_column_letter(offset)].width = max(12, min(24, len(text) + 3))

    style_data_area(ws, 6, 60, 2, 36)

    ws.freeze_panes = "B6"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 90

    try:
        wb.save(source)
        print(source.resolve())
    except PermissionError:
        wb.save(OUTPUT)
        print(OUTPUT.resolve())


if __name__ == "__main__":
    main()
