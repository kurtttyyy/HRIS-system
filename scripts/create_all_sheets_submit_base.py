from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from generate_employee_dashboard_excel import employees, headers


def thin(color="CBD5E1"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_nav(ws, active):
    items = [("Cover", "A3"), ("Dashboard", "A5"), ("Summary", "A7"), ("Data Entry", "A9"), ("Employee", "A11")]
    for text, cell_ref in items:
        target = "Employee Data Sheet" if text == "Data Entry" else text
        sheet_ref = f"'{target}'" if " " in target else target
        cell = ws[cell_ref]
        cell.value = text
        cell.hyperlink = f"#{sheet_ref}!A1"
        cell.fill = PatternFill("solid", fgColor="0F766E" if target == active else "075985")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin("0E7490")
    ws.column_dimensions["A"].width = 14


def base_sheet(ws, active):
    ws.sheet_view.showGridLines = False
    style_nav(ws, active)
    for col in range(2, 18):
        ws.column_dimensions[chr(64 + col)].width = 16
    for row in range(1, 36):
        ws.row_dimensions[row].height = 24


def make_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    base_sheet(ws, "Cover")
    for row in ws["B2:P30"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FBFAF5")
            cell.border = thin("F5E7BA")
    ws.merge_cells("B4:P6")
    ws["B4"] = "NORTHEASTERN COLLEGE"
    ws["B4"].font = Font(name="Georgia", bold=True, size=28, color="003D2B")
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B9:P12")
    ws["B9"] = "HUMAN RESOURCE"
    ws["B9"].font = Font(name="Georgia", bold=True, size=42, color="003D2B")
    ws["B9"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B13:P15")
    ws["B13"] = "D E P A R T M E N T"
    ws["B13"].font = Font(bold=True, size=24, color="003D2B")
    ws["B13"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B24:P27")
    ws["B24"] = "Empowering People. Building a Better Future."
    ws["B24"].fill = PatternFill("solid", fgColor="003D2B")
    ws["B24"].font = Font(name="Georgia", italic=True, size=22, color="F4E4B1")
    ws["B24"].alignment = Alignment(horizontal="center", vertical="center")


def make_employee(wb):
    ws = wb.create_sheet("Employee")
    base_sheet(ws, "Employee")
    start = 5
    for col, header in enumerate(headers, 1):
        c = ws.cell(start, col, header)
        c.fill = PatternFill("solid", fgColor="075985")
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.border = thin("0E7490")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, employee in enumerate(employees, start + 1):
        for col, header in enumerate(headers, 1):
            c = ws.cell(row_index, col, employee[header])
            c.border = thin()
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A6"


def make_data_entry(wb):
    ws = wb.create_sheet("Employee Data Sheet")
    base_sheet(ws, "Employee Data Sheet")
    ws.merge_cells("B2:P3")
    ws["B2"] = "Employee Data Sheet"
    ws["B2"].font = Font(bold=True, size=22)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    groups = [
        ("B5:E17", "Personal Information", "FFE4E6", [("Employee ID", "D6"), ("Name", "D7"), ("Age", "D8"), ("Gender", "D9"), ("Date of Birth", "D10"), ("Marital Status", "D11"), ("Nationality", "D12"), ("NID/Passport", "D13"), ("Living Area", "D14"), ("Email", "D15")]),
        ("F5:I17", "Job Information", "FEF3C7", [("Employment Type", "H6"), ("Designation", "H7"), ("Department", "H8"), ("Section", "H9"), ("Joining Date", "H10"), ("Reporting Boss", "H11"), ("Status", "H12")]),
        ("J5:M17", "Bank & Education", "E0F2FE", [("Bank Name", "L6"), ("Account No", "L7"), ("Payment Method", "L8"), ("Highest Education", "L11"), ("Top Training 1", "L12"), ("Top Training 2", "L13")]),
        ("N5:P17", "Salary Details", "DCFCE7", [("Joining Salary", "P6"), ("Current Salary", "P7"), ("Last Increment", "P8"), ("Performance Score", "P9"), ("Best Achievement", "P11")]),
    ]
    for ref, title, fill, fields in groups:
        for row in ws[ref]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.border = thin("E2E8F0")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        start_cell = ref.split(":")[0]
        ws[start_cell] = title
        ws[start_cell].font = Font(bold=True, color="075985", size=13)
        ws[start_cell].alignment = Alignment(horizontal="center")
        for label, cell_ref in fields:
            row = ws[cell_ref].row
            col = ws[cell_ref].column
            label_cell = ws.cell(row, col - 1, f"{label}:")
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            ws[cell_ref].fill = PatternFill("solid", fgColor="FFFFFF")
            ws[cell_ref].border = thin("94A3B8")

    ws["D6"] = '=CONCATENATE("A",TEXT(COUNTA(Employee!$A$6:$A$1000)+101,"000"))'
    ws["H12"] = "Active"
    ws["L8"] = "Bank Transfer"

    input_cells = {
        "Employee ID": "D6", "Name": "D7", "Age": "D8", "Gender": "D9", "Date of Birth": "D10",
        "Marital Status": "D11", "Nationality": "D12", "NID/Passport": "D13", "Living Area": "D14",
        "Email": "D15", "Employment Type": "H6", "Designation": "H7", "Department": "H8", "Section": "H9",
        "Joining Date": "H10", "Reporting Boss": "H11", "Status": "H12", "Bank Name": "L6",
        "Account No": "L7", "Payment Method": "L8", "Highest Education": "L11", "Top Training 1": "L12",
        "Top Training 2": "L13", "Joining Salary": "P6", "Current Salary": "P7", "Last Increment": "P8",
        "Performance Score": "P9", "Best Achievement": "P11",
    }
    ws.merge_cells("B20:P20")
    ws["B20"] = "Generated row for Employee page"
    ws["B20"].fill = PatternFill("solid", fgColor="075985")
    ws["B20"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["B20"].alignment = Alignment(horizontal="center")
    for col, header in enumerate(headers, 1):
        h = ws.cell(22, col, header)
        h.fill = PatternFill("solid", fgColor="075985")
        h.font = Font(bold=True, color="FFFFFF", size=9)
        h.border = thin("0E7490")
        h.alignment = Alignment(horizontal="center", wrap_text=True)
        c = ws.cell(23, col)
        if header == "Initials":
            c.value = '=IF(D7="","",UPPER(LEFT(D7,1)&IFERROR(MID(D7,FIND(" ",D7&" ")+1,1),"")))'
        elif header == "Service Age":
            c.value = '=IF(H10="","",ROUND((TODAY()-H10)/365.25,1))'
        elif header in input_cells:
            c.value = f"={input_cells[header]}"
        c.border = thin()
    ws.merge_cells("N18:P19")
    ws["N18"] = "Submit"
    ws["N18"].fill = PatternFill("solid", fgColor="0F766E")
    ws["N18"].font = Font(bold=True, color="FFFFFF", size=16)
    ws["N18"].alignment = Alignment(horizontal="center", vertical="center")


def make_summary(wb):
    ws = wb.create_sheet("Summary")
    base_sheet(ws, "Summary")
    ws.merge_cells("B2:J3")
    ws["B2"] = "HR Summary"
    ws["B2"].font = Font(bold=True, size=22)
    ws["B2"].alignment = Alignment(horizontal="center")
    rows = [("Active Employees", '=COUNTIF(Employee!$AC:$AC,"Active")'), ("Total Employees", "=COUNTA(Employee!$A:$A)-1"), ("Average Salary", '=AVERAGE(Employee!$Z:$Z)')]
    for i, (label, formula) in enumerate(rows, 6):
        ws.cell(i, 2, label)
        ws.cell(i, 3, formula)
        ws.cell(i, 2).font = Font(bold=True)


def make_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    base_sheet(ws, "Dashboard")
    ws.merge_cells("B2:P4")
    ws["B2"] = "Employee Dashboard"
    ws["B2"].font = Font(bold=True, size=24)
    ws["B2"].alignment = Alignment(horizontal="center")
    ws["B6"] = "Employee ID"
    ws["C6"] = "A101"
    ws["B8"] = "Name"
    ws["C8"] = '=IFERROR(INDEX(Employee!$B$6:$B$1000,MATCH($C$6,Employee!$A$6:$A$1000,0)),"")'
    ws["B9"] = "Department"
    ws["C9"] = '=IFERROR(INDEX(Employee!$N$6:$N$1000,MATCH($C$6,Employee!$A$6:$A$1000,0)),"")'
    ws["B10"] = "Performance Score"
    ws["C10"] = '=IFERROR(INDEX(Employee!$AB$6:$AB$1000,MATCH($C$6,Employee!$A$6:$A$1000,0)),"")'
    for cell in ["B6", "B8", "B9", "B10"]:
        ws[cell].font = Font(bold=True)


wb = Workbook()
make_cover(wb)
make_dashboard(wb)
make_summary(wb)
make_data_entry(wb)
make_employee(wb)
wb._sheets = [wb["Cover"], wb["Dashboard"], wb["Summary"], wb["Employee Data Sheet"], wb["Employee"]]
wb.active = 0
wb.save("employee_submit_all_sheets_base.xlsx")
print("employee_submit_all_sheets_base.xlsx")
