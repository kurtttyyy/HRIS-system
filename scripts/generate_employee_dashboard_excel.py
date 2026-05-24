from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFilter, ImageFont


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT = PROJECT_ROOT / "employee_dashboard_with_data_entry_fixed.xlsx"
LOGO_PATH = PROJECT_ROOT / "public" / "images" / "logo.webp"
COVER_IMAGE_PATH = PROJECT_ROOT / "public" / "images" / "employee_dashboard_cover.png"
WIDE_COVER_IMAGE_PATH = PROJECT_ROOT / "public" / "images" / "employee_dashboard_cover_wide.png"

BLUE = "075985"
TEAL = "2DD4BF"
GREEN = "DCFCE7"
PALE_BLUE = "E0F2FE"
PALE_YELLOW = "FEF3C7"
PALE_ROSE = "FFE4E6"
PALE_PURPLE = "EDE9FE"
GRAY = "F8FAFC"
DARK = "0F172A"
WHITE = "FFFFFF"
LINE = "CBD5E1"


employees = [
    {
        "Employee ID": "A101",
        "Name": "James Kooper",
        "Initials": "JK",
        "Age": 62,
        "Gender": "Male",
        "Date of Birth": date(1964, 6, 1),
        "Marital Status": "Married",
        "Nationality": "USA",
        "NID/Passport": "USA1100456",
        "Living Area": "Lane 1 - Zone Street",
        "Email": "james.kooper@example.com",
        "Employment Type": "Permanent",
        "Designation": "A.G.M",
        "Department": "Production",
        "Section": "Finishing",
        "Joining Date": date(1996, 6, 1),
        "Service Age": 29.0,
        "Reporting Boss": "Richard",
        "Bank Name": "Metro Bank",
        "Account No": "100245781",
        "Payment Method": "Bank Transfer",
        "Highest Education": "Bachelor of Science",
        "Top Training 1": "Brainstorming techniques",
        "Top Training 2": "Managing feedback",
        "Joining Salary": 20500,
        "Current Salary": 66713,
        "Last Increment": 2000,
        "Performance Score": 8,
        "Status": "Active",
        "Best Achievement": "Best Employee Award 2021",
    },
    {
        "Employee ID": "A102",
        "Name": "Maria Santos",
        "Initials": "MS",
        "Age": 35,
        "Gender": "Female",
        "Date of Birth": date(1991, 3, 19),
        "Marital Status": "Single",
        "Nationality": "Philippines",
        "NID/Passport": "PH778120",
        "Living Area": "Greenfield Residence",
        "Email": "maria.santos@example.com",
        "Employment Type": "Permanent",
        "Designation": "HR Officer",
        "Department": "HR",
        "Section": "Recruitment",
        "Joining Date": date(2018, 2, 10),
        "Service Age": 8.3,
        "Reporting Boss": "Anika",
        "Bank Name": "Union Bank",
        "Account No": "100245782",
        "Payment Method": "Bank Transfer",
        "Highest Education": "MBA",
        "Top Training 1": "Talent acquisition",
        "Top Training 2": "Labor compliance",
        "Joining Salary": 24500,
        "Current Salary": 52000,
        "Last Increment": 3500,
        "Performance Score": 9,
        "Status": "Active",
        "Best Achievement": "Reduced hiring cycle",
    },
    {
        "Employee ID": "A103",
        "Name": "David Miller",
        "Initials": "DM",
        "Age": 41,
        "Gender": "Male",
        "Date of Birth": date(1985, 11, 8),
        "Marital Status": "Married",
        "Nationality": "Canada",
        "NID/Passport": "CA992010",
        "Living Area": "Lakeview Avenue",
        "Email": "david.miller@example.com",
        "Employment Type": "Permanent",
        "Designation": "QA Lead",
        "Department": "Quality",
        "Section": "Inspection",
        "Joining Date": date(2014, 9, 5),
        "Service Age": 11.7,
        "Reporting Boss": "James Kooper",
        "Bank Name": "City Bank",
        "Account No": "100245783",
        "Payment Method": "Bank Transfer",
        "Highest Education": "BSc Engineering",
        "Top Training 1": "Root cause analysis",
        "Top Training 2": "ISO audit",
        "Joining Salary": 28000,
        "Current Salary": 61000,
        "Last Increment": 4500,
        "Performance Score": 7,
        "Status": "Active",
        "Best Achievement": "Zero defect month",
    },
    {
        "Employee ID": "A104",
        "Name": "William Chen",
        "Initials": "WC",
        "Age": 29,
        "Gender": "Male",
        "Date of Birth": date(1997, 1, 24),
        "Marital Status": "Single",
        "Nationality": "Singapore",
        "NID/Passport": "SG450012",
        "Living Area": "Riverside Road",
        "Email": "william.chen@example.com",
        "Employment Type": "Contract",
        "Designation": "IT Analyst",
        "Department": "IT",
        "Section": "Systems",
        "Joining Date": date(2022, 1, 3),
        "Service Age": 4.4,
        "Reporting Boss": "Nora",
        "Bank Name": "Digital Trust",
        "Account No": "100245784",
        "Payment Method": "Bank Transfer",
        "Highest Education": "BSc Computer Science",
        "Top Training 1": "Cybersecurity basics",
        "Top Training 2": "Cloud administration",
        "Joining Salary": 32000,
        "Current Salary": 47000,
        "Last Increment": 4000,
        "Performance Score": 8,
        "Status": "Active",
        "Best Achievement": "Server migration",
    },
    {
        "Employee ID": "A105",
        "Name": "Richard Adams",
        "Initials": "RA",
        "Age": 48,
        "Gender": "Male",
        "Date of Birth": date(1978, 8, 14),
        "Marital Status": "Married",
        "Nationality": "USA",
        "NID/Passport": "USA445879",
        "Living Area": "North Star Court",
        "Email": "richard.adams@example.com",
        "Employment Type": "Permanent",
        "Designation": "Production Manager",
        "Department": "Production",
        "Section": "Assembly",
        "Joining Date": date(2007, 4, 17),
        "Service Age": 19.1,
        "Reporting Boss": "James Kooper",
        "Bank Name": "Metro Bank",
        "Account No": "100245785",
        "Payment Method": "Bank Transfer",
        "Highest Education": "Industrial Management",
        "Top Training 1": "Lean manufacturing",
        "Top Training 2": "Team leadership",
        "Joining Salary": 30000,
        "Current Salary": 74000,
        "Last Increment": 6000,
        "Performance Score": 9,
        "Status": "Active",
        "Best Achievement": "Output growth 18%",
    },
    {
        "Employee ID": "A106",
        "Name": "Joseph Cruz",
        "Initials": "JC",
        "Age": 33,
        "Gender": "Male",
        "Date of Birth": date(1993, 12, 2),
        "Marital Status": "Single",
        "Nationality": "Philippines",
        "NID/Passport": "PH772219",
        "Living Area": "South District",
        "Email": "joseph.cruz@example.com",
        "Employment Type": "Temporary",
        "Designation": "Accountant",
        "Department": "Accounts",
        "Section": "Payroll",
        "Joining Date": date(2020, 6, 18),
        "Service Age": 5.9,
        "Reporting Boss": "Lara",
        "Bank Name": "Union Bank",
        "Account No": "100245786",
        "Payment Method": "Bank Transfer",
        "Highest Education": "Bachelor of Accountancy",
        "Top Training 1": "Tax update",
        "Top Training 2": "Payroll controls",
        "Joining Salary": 26000,
        "Current Salary": 43000,
        "Last Increment": 3000,
        "Performance Score": 7,
        "Status": "Inactive",
        "Best Achievement": "Payroll cleanup",
    },
]


headers = list(employees[0].keys())
col_index = {header: i + 1 for i, header in enumerate(headers)}


def lookup_formula(employee_id_cell, header):
    col = col_index[header]
    letter = get_column_letter(col)
    return (
        f'=IFERROR(INDEX(Employee!${letter}$6:${letter}$105,'
        f"MATCH({employee_id_cell},Employee!$A$6:$A$105,0)),\"\")"
    )


def style_sheet(ws, show_grid=False):
    ws.sheet_view.showGridLines = show_grid
    ws.freeze_panes = "A1"


def thin_border(color=LINE):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def set_title(cell, text, size=24, color=WHITE):
    cell.value = text
    cell.font = Font(name="Aptos Display", bold=True, size=size, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def merge_card(ws, ref, fill, border_color=LINE):
    rows = ws[ref]
    for row in rows:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = thin_border(border_color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_nav(ws, active):
    nav_items = [("Cover", "A3"), ("Dashboard", "A5"), ("Summary", "A7"), ("Data Entry", "A9"), ("Employee", "A11")]
    for name, cell_ref in nav_items:
        cell = ws[cell_ref]
        target = "Employee Data Sheet" if name == "Data Entry" else name
        cell.value = name
        sheet_ref = f"'{target}'" if " " in target else target
        cell.hyperlink = f"#{sheet_ref}!A1"
        cell.style = "Hyperlink"
        cell.font = Font(bold=True, color=WHITE if target == active else "D1FAE5", size=10)
        cell.fill = PatternFill("solid", fgColor="0F766E" if target == active else BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border("0E7490")
        ws.row_dimensions[cell.row].height = 26
    ws.column_dimensions["A"].width = 14


def font(size, bold=False, italic=False):
    candidates = [
        "C:/Windows/Fonts/georgiab.ttf" if bold else "C:/Windows/Fonts/georgia.ttf",
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
    ]
    if italic:
        candidates.insert(0, "C:/Windows/Fonts/georgiai.ttf")
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()


def centered_text(draw, box, text, text_font, fill, spacing=8):
    lines = text.split("\n")
    line_heights = []
    widths = []
    for line in lines:
        bbox = draw.textbbox((0, 0), line, font=text_font)
        widths.append(bbox[2] - bbox[0])
        line_heights.append(bbox[3] - bbox[1])
    total_height = sum(line_heights) + spacing * (len(lines) - 1)
    x1, y1, x2, y2 = box
    y = y1 + ((y2 - y1) - total_height) / 2
    for line, width, height in zip(lines, widths, line_heights):
        x = x1 + ((x2 - x1) - width) / 2
        draw.text((x, y), line, font=text_font, fill=fill)
        y += height + spacing


def create_cover_image():
    COVER_IMAGE_PATH.parent.mkdir(parents=True, exist_ok=True)
    width, height = 1100, 1500
    dark_green = (0, 55, 38)
    green = (13, 95, 63)
    light_green = (137, 160, 109)
    gold = (201, 154, 49)
    pale_gold = (249, 221, 139)
    cream = (252, 247, 229)
    deep = (8, 52, 37)

    img = Image.new("RGB", (width, height), cream)
    pixels = img.load()
    for y in range(height):
        for x in range(width):
            dx = x / width
            dy = y / height
            glow = int(28 * (1 - ((dx - 0.52) ** 2 + (dy - 0.35) ** 2)))
            shade = int(18 * dy)
            r = max(0, min(255, cream[0] + glow - shade))
            g = max(0, min(255, cream[1] + glow - shade))
            b = max(0, min(255, cream[2] + glow - shade))
            pixels[x, y] = (r, g, b)

    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    draw.polygon([(0, 0), (430, 0), (300, 95), (95, 170), (0, 300)], fill=dark_green + (255,))
    draw.polygon([(0, 300), (95, 170), (300, 95), (430, 0), (475, 0), (320, 135), (110, 225), (0, 385)], fill=gold + (230,))
    for i in range(8):
        draw.arc((-250 + i * 18, -115 + i * 9, 580 + i * 18, 270 + i * 9), 185, 350, fill=(154, 119, 35, 95), width=1)

    draw.polygon([(0, 1120), (180, 1190), (360, 1225), (575, 1258), (760, 1260), (960, 1185), (1100, 1040), (1100, 1500), (0, 1500)], fill=dark_green + (255,))
    draw.polygon([(0, 1090), (260, 1170), (515, 1200), (720, 1180), (900, 1115), (1100, 960), (1100, 1060), (935, 1170), (720, 1250), (500, 1275), (260, 1248), (0, 1160)], fill=light_green + (125,))
    draw.polygon([(255, 1265), (470, 1207), (705, 1192), (940, 1168), (1100, 1108), (1100, 1168), (920, 1220), (708, 1245), (485, 1252), (265, 1310)], fill=gold + (235,))

    for i in range(14):
        x = 660 + i * 26
        for j in range(12):
            radius = 2 + j // 4
            draw.ellipse((x, 58 + j * 22, x + radius * 2, 58 + j * 22 + radius * 2), fill=(203, 159, 65, 75))

    img = Image.alpha_composite(img.convert("RGBA"), overlay)
    draw = ImageDraw.Draw(img)

    draw.rounded_rectangle((38, 38, width - 38, height - 38), radius=0, outline=gold, width=4)

    if LOGO_PATH.exists():
        logo = Image.open(LOGO_PATH).convert("RGBA")
        logo.thumbnail((280, 280), Image.Resampling.LANCZOS)
        shadow = Image.new("RGBA", logo.size, (0, 0, 0, 0))
        shadow.alpha_composite(logo)
        shadow = shadow.filter(ImageFilter.GaussianBlur(8))
        lx = (width - logo.width) // 2
        ly = 88
        img.alpha_composite(shadow, (lx + 4, ly + 8))
        img.alpha_composite(logo, (lx, ly))

    centered_text(draw, (140, 386, width - 140, 440), "NORTHEASTERN COLLEGE", font(38, bold=True), deep)
    draw.line((350, 455, 445, 455), fill=gold, width=3)
    draw.line((655, 455, 750, 455), fill=gold, width=3)
    centered_text(draw, (275, 445, width - 275, 482), "KNOWLEDGE IS POWER", font(18), (132, 95, 22))

    centered_text(draw, (60, 550, width - 60, 645), "HUMAN RESOURCE", font(74, bold=True), deep)
    draw.line((185, 692, 285, 692), fill=gold, width=5)
    draw.line((815, 692, 915, 692), fill=gold, width=5)
    centered_text(draw, (290, 655, 810, 735), "D E P A R T M E N T", font(46, bold=True), deep)

    draw.line((410, 790, 690, 790), fill=gold, width=2)
    draw.ellipse((534, 778, 566, 810), fill=gold)
    draw.ellipse((500, 784, 526, 802), fill=gold)
    draw.ellipse((574, 784, 600, 802), fill=gold)

    draw.ellipse((450, 850, 650, 1050), fill=dark_green, outline=gold, width=5)
    centered_text(draw, (450, 850, 650, 1050), "HR", font(60, bold=True), pale_gold)

    values = [
        ("OUR PEOPLE", "Our Greatest Strength"),
        ("RESPECT", "Value Every Individual"),
        ("INTEGRITY", "Uphold Trust and Ethics"),
        ("GROWTH", "Empower Potential"),
    ]
    centers = [205, 435, 665, 895]
    for idx, ((title, subtitle), cx) in enumerate(zip(values, centers)):
        if idx:
            draw.line((cx - 115, 1280, cx - 115, 1410), fill=gold, width=2)
        centered_text(draw, (cx - 95, 1294, cx + 95, 1332), title, font(22, bold=True), (255, 255, 255))
        centered_text(draw, (cx - 110, 1335, cx + 110, 1380), subtitle, font(18, italic=True), pale_gold)

    draw.line((140, 1455, 265, 1455), fill=gold, width=2)
    draw.line((835, 1455, 960, 1455), fill=gold, width=2)
    centered_text(draw, (260, 1428, 840, 1485), "Empowering People. Building a Better Future.", font(29, italic=True), pale_gold)

    img.convert("RGB").save(COVER_IMAGE_PATH, "PNG", quality=95)
    return COVER_IMAGE_PATH


def create_wide_cover_image():
    WIDE_COVER_IMAGE_PATH.parent.mkdir(parents=True, exist_ok=True)
    width, height = 1800, 900
    dark_green = (0, 55, 38)
    green = (12, 92, 62)
    light_green = (143, 162, 112)
    gold = (202, 154, 44)
    pale_gold = (249, 222, 141)
    cream = (252, 247, 229)
    deep = (7, 54, 38)

    img = Image.new("RGB", (width, height), cream)
    pixels = img.load()
    for y in range(height):
        for x in range(width):
            dx = x / width
            dy = y / height
            glow = int(34 * (1 - ((dx - 0.54) ** 2 + (dy - 0.38) ** 2)))
            shade = int(16 * dy)
            pixels[x, y] = (
                max(0, min(255, cream[0] + glow - shade)),
                max(0, min(255, cream[1] + glow - shade)),
                max(0, min(255, cream[2] + glow - shade)),
            )

    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    draw.polygon([(0, 0), (720, 0), (545, 85), (230, 150), (0, 330)], fill=dark_green + (255,))
    draw.polygon([(0, 330), (210, 160), (545, 85), (720, 0), (805, 0), (585, 125), (250, 210), (0, 440)], fill=gold + (228,))
    for i in range(10):
        draw.arc((-260 + i * 22, -180 + i * 10, 920 + i * 20, 260 + i * 9), 182, 350, fill=(142, 105, 27, 88), width=1)

    draw.polygon([(0, 710), (340, 775), (700, 795), (1040, 770), (1380, 700), (1800, 560), (1800, 900), (0, 900)], fill=dark_green + (255,))
    draw.polygon([(0, 675), (350, 745), (720, 758), (1070, 730), (1390, 660), (1800, 500), (1800, 600), (1420, 720), (1050, 800), (690, 827), (320, 805), (0, 740)], fill=light_green + (118,))
    draw.polygon([(500, 815), (810, 755), (1140, 728), (1490, 690), (1800, 625), (1800, 695), (1480, 760), (1125, 810), (780, 835), (500, 870)], fill=gold + (238,))

    for i in range(22):
        x = 1160 + i * 24
        for j in range(9):
            radius = 2 + j // 3
            draw.ellipse((x, 55 + j * 24, x + radius * 2, 55 + j * 24 + radius * 2), fill=(203, 159, 65, 70))

    img = Image.alpha_composite(img.convert("RGBA"), overlay)
    draw = ImageDraw.Draw(img)
    draw.rounded_rectangle((36, 36, width - 36, height - 36), radius=0, outline=gold, width=4)

    if LOGO_PATH.exists():
        logo = Image.open(LOGO_PATH).convert("RGBA")
        logo.thumbnail((205, 205), Image.Resampling.LANCZOS)
        shadow = Image.new("RGBA", logo.size, (0, 0, 0, 0))
        shadow.alpha_composite(logo)
        shadow = shadow.filter(ImageFilter.GaussianBlur(7))
        lx = (width - logo.width) // 2
        ly = 64
        img.alpha_composite(shadow, (lx + 4, ly + 7))
        img.alpha_composite(logo, (lx, ly))

    centered_text(draw, (360, 262, width - 360, 315), "NORTHEASTERN COLLEGE", font(38, bold=True), deep)
    draw.line((670, 327, 812, 327), fill=gold, width=3)
    draw.line((988, 327, 1130, 327), fill=gold, width=3)
    centered_text(draw, (720, 318, 1080, 354), "KNOWLEDGE IS POWER", font(17), (132, 95, 22))

    centered_text(draw, (140, 402, width - 140, 505), "HUMAN RESOURCE", font(78, bold=True), deep)
    draw.line((420, 555, 560, 555), fill=gold, width=5)
    draw.line((1240, 555, 1380, 555), fill=gold, width=5)
    centered_text(draw, (570, 518, 1230, 596), "D E P A R T M E N T", font(48, bold=True), deep)

    draw.line((775, 642, 1025, 642), fill=gold, width=2)
    draw.ellipse((884, 629, 916, 661), fill=gold)
    draw.ellipse((846, 635, 872, 655), fill=gold)
    draw.ellipse((928, 635, 954, 655), fill=gold)

    draw.ellipse((815, 680, 985, 850), fill=dark_green, outline=gold, width=5)
    centered_text(draw, (815, 680, 985, 850), "HR", font(56, bold=True), pale_gold)

    values = [
        ("OUR PEOPLE", "Our Greatest Strength"),
        ("RESPECT", "Value Every Individual"),
        ("INTEGRITY", "Uphold Trust and Ethics"),
        ("GROWTH", "Empower Potential"),
    ]
    centers = [305, 650, 1150, 1495]
    for idx, ((title, subtitle), cx) in enumerate(zip(values, centers)):
        draw.rounded_rectangle((cx - 160, 754, cx + 160, 840), radius=18, fill=(0, 70, 48, 220), outline=gold, width=2)
        centered_text(draw, (cx - 142, 766, cx + 142, 795), title, font(20, bold=True), (255, 255, 255))
        centered_text(draw, (cx - 142, 801, cx + 142, 828), subtitle, font(13, italic=True), pale_gold)

    draw.line((500, 865, 670, 865), fill=gold, width=2)
    draw.line((1130, 865, 1300, 865), fill=gold, width=2)
    centered_text(draw, (660, 842, 1140, 892), "Empowering People. Building a Better Future.", font(25, italic=True), pale_gold)

    img.convert("RGB").save(WIDE_COVER_IMAGE_PATH, "PNG", quality=95)
    return WIDE_COVER_IMAGE_PATH


def make_employee_sheet(wb):
    ws = wb.create_sheet("Employee")
    style_sheet(ws, True)
    add_nav(ws, "Employee")

    ws.merge_cells("C2:AB2")
    set_title(ws["C2"], "Employee Data Registry", 22, DARK)
    ws["C3"] = "Edit or add employee records here. Dashboard and Summary update from this table."
    ws["C3"].font = Font(italic=True, color="475569")

    start_row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border("0E7490")

    for row_index, employee in enumerate(employees, start_row + 1):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row_index, col, employee[header])
            cell.fill = PatternFill("solid", fgColor=GRAY if row_index % 2 == 0 else WHITE)
            cell.border = thin_border("E2E8F0")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if header in {"Joining Salary", "Current Salary", "Last Increment"}:
                cell.number_format = '#,##0'
            if isinstance(employee[header], date):
                cell.number_format = "d-mmm-yyyy"

    table_ref = f"A{start_row}:AD{start_row + len(employees)}"
    tab = Table(displayName="EmployeeTable", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    widths = {
        "A": 12,
        "B": 20,
        "C": 10,
        "D": 8,
        "E": 12,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 16,
        "J": 22,
        "K": 26,
        "L": 16,
        "M": 18,
        "N": 16,
        "O": 16,
        "P": 14,
        "Q": 12,
        "R": 18,
        "S": 16,
        "T": 15,
        "U": 16,
        "V": 24,
        "W": 24,
        "X": 24,
        "Y": 14,
        "Z": 14,
        "AA": 14,
        "AB": 16,
        "AC": 12,
        "AD": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = table_ref

    return ws


def make_cover_sheet(wb):
    ws = wb.active
    ws.title = "Cover"
    style_sheet(ws)

    for col in range(1, 28):
        ws.column_dimensions[get_column_letter(col)].width = 10
    for row in range(1, 42):
        ws.row_dimensions[row].height = 20

    for row in ws["A1:AA41"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FBFAF5")
            cell.border = Border()

    cover = XLImage(str(create_wide_cover_image()))
    cover.width = 1500
    cover.height = 750
    ws.add_image(cover, "B2")

    ws.merge_cells("Y3:Z4")
    ws["Y3"] = "Open\nDashboard"
    ws["Y3"].hyperlink = "#Dashboard!A1"
    ws["Y3"].font = Font(bold=True, color=WHITE, size=10)
    ws["Y3"].fill = PatternFill("solid", fgColor="003D2B")
    ws["Y3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def make_employee_data_sheet(wb):
    ws = wb.create_sheet("Employee Data Sheet")
    style_sheet(ws)
    add_nav(ws, "Employee Data Sheet")

    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for row in range(1, 42):
        ws.row_dimensions[row].height = 24

    for row in ws["B2:P38"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="F8FAFC")
            cell.border = Border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells("B2:P3")
    ws["B2"] = "Employee Data Sheet"
    ws["B2"].font = Font(name="Aptos Display", bold=True, size=24, color=DARK)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B4:P4")
    ws["B4"] = "Admin entry form. Fill the fields below, then copy the generated row at the bottom into the next blank row of the Employee page."
    ws["B4"].font = Font(italic=True, color="475569")
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

    input_cells = {
        "Employee ID": "D6",
        "Name": "D7",
        "Age": "D8",
        "Gender": "D9",
        "Date of Birth": "D10",
        "Marital Status": "D11",
        "Nationality": "D12",
        "NID/Passport": "D13",
        "Living Area": "D14",
        "Email": "D15",
        "Employment Type": "H6",
        "Designation": "H7",
        "Department": "H8",
        "Section": "H9",
        "Joining Date": "H10",
        "Reporting Boss": "H11",
        "Status": "H12",
        "Bank Name": "L6",
        "Account No": "L7",
        "Payment Method": "L8",
        "Highest Education": "L11",
        "Top Training 1": "L12",
        "Top Training 2": "L13",
        "Joining Salary": "P6",
        "Current Salary": "P7",
        "Last Increment": "P8",
        "Performance Score": "P9",
        "Best Achievement": "P11",
    }

    groups = [
        ("B5:E17", "Personal Information", PALE_ROSE, [
            ("Employee ID", "D6"),
            ("Name", "D7"),
            ("Age", "D8"),
            ("Gender", "D9"),
            ("Date of Birth", "D10"),
            ("Marital Status", "D11"),
            ("Nationality", "D12"),
            ("NID/Passport", "D13"),
            ("Living Area", "D14"),
            ("Email", "D15"),
        ]),
        ("F5:I17", "Job Information", PALE_YELLOW, [
            ("Employment Type", "H6"),
            ("Designation", "H7"),
            ("Department", "H8"),
            ("Section", "H9"),
            ("Joining Date", "H10"),
            ("Reporting Boss", "H11"),
            ("Status", "H12"),
        ]),
        ("J5:M17", "Bank & Education", PALE_BLUE, [
            ("Bank Name", "L6"),
            ("Account No", "L7"),
            ("Payment Method", "L8"),
            ("Highest Education", "L11"),
            ("Top Training 1", "L12"),
            ("Top Training 2", "L13"),
        ]),
        ("N5:P17", "Salary Details", GREEN, [
            ("Joining Salary", "P6"),
            ("Current Salary", "P7"),
            ("Last Increment", "P8"),
            ("Performance Score", "P9"),
            ("Best Achievement", "P11"),
        ]),
    ]

    for ref, title, fill, fields in groups:
        merge_card(ws, ref, fill, "E2E8F0")
        start_cell, end_cell = ref.split(":")
        start_col = ws[start_cell].column
        start_row = ws[start_cell].row
        end_col = ws[end_cell].column
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        ws.cell(start_row, start_col, title)
        ws.cell(start_row, start_col).font = Font(bold=True, size=13, color=BLUE)
        ws.cell(start_row, start_col).alignment = Alignment(horizontal="center", vertical="center")

        for label, cell_ref in fields:
            row = ws[cell_ref].row
            label_cell = ws.cell(row, ws[cell_ref].column - 1, f"{label}:")
            value_cell = ws[cell_ref]
            label_cell.font = Font(bold=True, color=DARK, size=10)
            label_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            value_cell.fill = PatternFill("solid", fgColor=WHITE)
            value_cell.border = thin_border("94A3B8")
            value_cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws["D6"] = '=CONCATENATE("A",TEXT(COUNTA(Employee!$A$6:$A$105)+101,"000"))'
    ws["H12"] = "Active"
    ws["L8"] = "Bank Transfer"

    list_validations = {
        "D9": '"Male,Female,Other"',
        "D11": '"Single,Married,Widowed,Separated"',
        "H6": '"Permanent,Contract,Temporary,Intern"',
        "H8": '"Production,Quality,HR,IT,Accounts"',
        "H12": '"Active,Inactive"',
        "L8": '"Bank Transfer,Cash,Check"',
    }
    for cell_ref, formula in list_validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws[cell_ref])

    for cell_ref in ["D10", "H10"]:
        ws[cell_ref].number_format = "d-mmm-yyyy"
    for cell_ref in ["P6", "P7", "P8"]:
        ws[cell_ref].number_format = '#,##0'

    ws.merge_cells("B20:P20")
    ws["B20"] = "Generated row for Employee page"
    ws["B20"].font = Font(bold=True, size=14, color=WHITE)
    ws["B20"].fill = PatternFill("solid", fgColor=BLUE)
    ws["B20"].alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        header_cell = ws.cell(22, col, header)
        header_cell.fill = PatternFill("solid", fgColor=BLUE)
        header_cell.font = Font(bold=True, color=WHITE, size=9)
        header_cell.border = thin_border("0E7490")
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        source = input_cells.get(header)
        value_cell = ws.cell(23, col)
        if header == "Initials":
            value_cell.value = '=IF(D7="","",UPPER(LEFT(D7,1)&IFERROR(MID(D7,FIND(" ",D7&" ")+1,1),"")))'
        elif header == "Service Age":
            value_cell.value = '=IF(H10="","",ROUND((TODAY()-H10)/365.25,1))'
        elif source:
            value_cell.value = f"={source}"
        else:
            value_cell.value = ""
        value_cell.fill = PatternFill("solid", fgColor=WHITE)
        value_cell.border = thin_border("CBD5E1")
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells("B26:P28")
    ws["B26"] = "Save to Employee page: copy row 23 from this sheet, go to Employee page, select the next blank row under the table, then paste values. Dashboard and Summary will update after recalculation."
    ws["B26"].font = Font(color="334155", italic=True)
    ws["B26"].fill = PatternFill("solid", fgColor="E0F2FE")
    ws["B26"].border = thin_border("7DD3FC")
    ws["B26"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("B30:D31")
    ws["B30"] = "Go to Employee Page"
    ws["B30"].hyperlink = "#Employee!A1"
    ws["B30"].font = Font(bold=True, color=WHITE)
    ws["B30"].fill = PatternFill("solid", fgColor="0F766E")
    ws["B30"].alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "B22"


def make_dashboard_sheet(wb):
    ws = wb.create_sheet("Dashboard")
    style_sheet(ws)
    add_nav(ws, "Dashboard")

    widths = {
        "A": 13,
        "B": 16,
        "C": 18,
        "D": 16,
        "E": 18,
        "F": 16,
        "G": 18,
        "H": 16,
        "I": 18,
        "J": 16,
        "K": 18,
        "L": 16,
        "M": 18,
        "N": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, 45):
        ws.row_dimensions[row].height = 24
    for row in [2, 3, 4, 5, 7, 8, 17, 18, 27]:
        ws.row_dimensions[row].height = 30

    for row in ws["B1:N42"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="F8FAFC")
            cell.border = Border()

    ws.merge_cells("B2:C5")
    merge_card(ws, "B2:C5", "DBEAFE", "93C5FD")
    ws["B2"] = lookup_formula("$N$3", "Initials")
    ws["B2"].font = Font(bold=True, size=30, color=BLUE)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("D2:I2")
    ws["D2"] = lookup_formula("$N$3", "Name")
    ws["D2"].font = Font(name="Aptos Display", bold=True, size=24, color=DARK)
    ws["D2"].alignment = Alignment(vertical="center")

    ws.merge_cells("D3:I3")
    ws["D3"] = '=CONCATENATE(IFERROR(INDEX(Employee!$M$6:$M$105,MATCH($N$3,Employee!$A$6:$A$105,0)),""),"  |  ",IFERROR(INDEX(Employee!$N$6:$N$105,MATCH($N$3,Employee!$A$6:$A$105,0)),""))'
    ws["D3"].font = Font(size=12, color="334155")

    ws.merge_cells("D4:I4")
    ws["D4"] = '=CONCATENATE("Achievement: ",IFERROR(INDEX(Employee!$AD$6:$AD$105,MATCH($N$3,Employee!$A$6:$A$105,0)),""))'
    ws["D4"].font = Font(size=11, color="475569")

    ws.merge_cells("D5:I5")
    ws["D5"] = '=CONCATENATE("Email: ",IFERROR(INDEX(Employee!$K$6:$K$105,MATCH($N$3,Employee!$A$6:$A$105,0)),""))'
    ws["D5"].font = Font(size=11, color="475569")

    merge_card(ws, "J2:L5", "CCFBF1", "5EEAD4")
    ws.merge_cells("J2:L2")
    ws["J2"] = "Performance Score"
    ws["J2"].font = Font(bold=True, size=14, color=BLUE)
    ws["J2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("J3:L5")
    ws["J3"] = lookup_formula("$N$3", "Performance Score")
    ws["J3"].font = Font(bold=True, size=34, color="064E3B")
    ws["J3"].alignment = Alignment(horizontal="center", vertical="center")

    ws["N2"] = "Employee ID"
    ws["N2"].font = Font(bold=True, color=WHITE)
    ws["N2"].fill = PatternFill("solid", fgColor=BLUE)
    ws["N2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["N3"] = "A101"
    ws["N3"].font = Font(bold=True, size=12)
    ws["N3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["N3"].fill = PatternFill("solid", fgColor=WHITE)
    ws["N3"].border = thin_border(BLUE)
    dv = DataValidation(type="list", formula1="=Employee!$A$6:$A$105", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["N3"])

    def add_panel(title, start_row, start_col, end_col, fill, fields):
        end_row = start_row + len(fields) + 2
        merge_card(
            ws,
            f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}",
            fill,
            "E2E8F0",
        )
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        title_cell = ws.cell(start_row, start_col, title)
        title_cell.font = Font(bold=True, size=13, color=BLUE)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        label_end = start_col + 1
        value_start = start_col + 2
        for offset, field in enumerate(fields, start_row + 2):
            ws.merge_cells(start_row=offset, start_column=start_col, end_row=offset, end_column=label_end)
            ws.merge_cells(start_row=offset, start_column=value_start, end_row=offset, end_column=end_col)
            label = ws.cell(offset, start_col, f"{field}:")
            value = ws.cell(offset, value_start, lookup_formula("$N$3", field))
            label.font = Font(bold=True, color=DARK, size=10)
            value.font = Font(color="1E293B", size=10)
            label.alignment = Alignment(vertical="center", wrap_text=True)
            value.alignment = Alignment(vertical="center", wrap_text=True)
            if field in {"Joining Salary", "Current Salary", "Last Increment"}:
                value.number_format = '#,##0'
            if field == "Date of Birth":
                value.number_format = "d-mmm-yyyy"

    add_panel(
        "Personal Detail",
        7,
        2,
        5,
        PALE_ROSE,
        ["Date of Birth", "Age", "Gender", "Marital Status", "Nationality", "NID/Passport", "Living Area"],
    )
    add_panel(
        "Job Detail",
        7,
        6,
        9,
        PALE_YELLOW,
        ["Employment Type", "Designation", "Service Age", "Department", "Section", "Reporting Boss", "Status"],
    )
    add_panel(
        "Bank Detail",
        7,
        10,
        13,
        PALE_BLUE,
        ["Bank Name", "Account No", "Payment Method"],
    )
    add_panel(
        "Educational Background",
        18,
        2,
        6,
        PALE_PURPLE,
        ["Highest Education", "Top Training 1", "Top Training 2"],
    )
    add_panel(
        "Salary Detail",
        18,
        7,
        13,
        GREEN,
        ["Joining Salary", "Current Salary", "Last Increment", "Performance Score"],
    )

    merge_card(ws, "B28:E40", WHITE)
    chart = DoughnutChart()
    data = Reference(wb["Summary"], min_col=3, min_row=4, max_row=6)
    labels = Reference(wb["Summary"], min_col=2, min_row=4, max_row=6)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    chart.title = "Gender Mix"
    chart.holeSize = 55
    chart.height = 8
    chart.width = 8.5
    ws.add_chart(chart, "B28")

    merge_card(ws, "F28:I40", WHITE)
    bar = BarChart()
    data = Reference(wb["Summary"], min_col=6, min_row=3, max_row=8)
    labels = Reference(wb["Summary"], min_col=5, min_row=4, max_row=8)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = "Manpower by Department"
    bar.height = 8
    bar.width = 8.5
    ws.add_chart(bar, "F28")

    merge_card(ws, "J28:M40", WHITE)
    line = LineChart()
    data = Reference(wb["Summary"], min_col=9, min_row=3, max_row=9)
    labels = Reference(wb["Summary"], min_col=8, min_row=4, max_row=9)
    line.add_data(data, titles_from_data=True)
    line.set_categories(labels)
    line.title = "Join by Year"
    line.height = 8
    line.width = 8.5
    ws.add_chart(line, "J28")


def make_summary_sheet(wb):
    ws = wb.create_sheet("Summary")
    style_sheet(ws)
    add_nav(ws, "Summary")

    ws.merge_cells("B2:J2")
    set_title(ws["B2"], "HR Summary", 22, DARK)

    ws["B4"], ws["C4"] = "Male", '=COUNTIF(Employee!$E$6:$E$105,"Male")'
    ws["B5"], ws["C5"] = "Female", '=COUNTIF(Employee!$E$6:$E$105,"Female")'
    ws["B6"], ws["C6"] = "Other", '=COUNTA(Employee!$A$6:$A$105)-SUM(C4:C5)'
    ws["B8"], ws["C8"] = "Active Employees", '=COUNTIF(Employee!$AC$6:$AC$105,"Active")'
    ws["B9"], ws["C9"] = "Average Salary", '=AVERAGEIF(Employee!$AC$6:$AC$105,"Active",Employee!$Z$6:$Z$105)'
    ws["B10"], ws["C10"] = "Average Performance", '=AVERAGE(Employee!$AB$6:$AB$105)'

    departments = ["Production", "Quality", "HR", "IT", "Accounts"]
    ws["E3"], ws["F3"] = "Department", "Headcount"
    for row, department in enumerate(departments, 4):
        ws.cell(row, 5, department)
        ws.cell(row, 6, f'=COUNTIF(Employee!$N$6:$N$105,E{row})')

    years = [1996, 2007, 2014, 2018, 2020, 2022]
    ws["H3"], ws["I3"] = "Year", "Join Count"
    for row, year in enumerate(years, 4):
        ws.cell(row, 8, year)
        ws.cell(row, 9, f'=SUMPRODUCT(--(YEAR(Employee!$P$6:$P$105)=H{row}))')

    ws["K3"], ws["L3"] = "Employee", "Performance"
    for row, employee in enumerate(employees, 4):
        ws.cell(row, 11, employee["Name"])
        ws.cell(row, 12, f'=IFERROR(INDEX(Employee!$AB$6:$AB$105,MATCH(K{row},Employee!$B$6:$B$105,0)),"")')

    for area in ["B3:C10", "E3:F8", "H3:I9", "K3:L9"]:
        for row in ws[area]:
            for cell in row:
                cell.border = thin_border()
                cell.alignment = Alignment(vertical="center")
                if cell.row == 3:
                    cell.fill = PatternFill("solid", fgColor=BLUE)
                    cell.font = Font(bold=True, color=WHITE)
                else:
                    cell.fill = PatternFill("solid", fgColor=GRAY if cell.row % 2 == 0 else WHITE)

    ws["C9"].number_format = '#,##0'
    ws.conditional_formatting.add("L4:L9", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color=TEAL))

    for col in ["B", "E", "H", "K"]:
        ws.column_dimensions[col].width = 22
    for col in ["C", "F", "I", "L"]:
        ws.column_dimensions[col].width = 16


def main():
    wb = Workbook()
    make_cover_sheet(wb)
    make_employee_sheet(wb)
    make_employee_data_sheet(wb)
    make_summary_sheet(wb)
    make_dashboard_sheet(wb)

    wb._sheets = [wb["Cover"], wb["Dashboard"], wb["Summary"], wb["Employee Data Sheet"], wb["Employee"]]
    wb.active = 0
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.zoomScale = 90

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
